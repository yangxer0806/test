# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os, re
from openpyxl import load_workbook, Workbook

dl = ['胸围', '肩宽']
dl2 = ['Bust', 'Shoulder']
ad = []
# 获取当前路径
cur_dir = os.path.dirname(os.path.abspath(__file__))
# 遍历Excel文件夹
for root, dirs, files in os.walk(cur_dir + '\\Excel'):
 for file in files:
  d = []
  wb = load_workbook(filename='Excel/' + file, data_only=True)
  sheet_name = wb.sheetnames[3]
  sheet_ranges = wb[sheet_name]
  # 取款号
  kh = sheet_ranges['C2'].value
  print(sheet_ranges['C2'].value)
  # 循环判断是否是要的度量方法
  for cell in sheet_ranges['B']:
   cv = cell.value
   ci = cell.row
   if cv != None:
    # 判断是否要的度量
    t = 0
    for i in dl:
     if re.match(i, cv) != None:
      xsv = sheet_ranges['D' + str(ci)].value
      sv = sheet_ranges['E' + str(ci)].value
      mv = sheet_ranges['F' + str(ci)].value
      lv = sheet_ranges['G' + str(ci)].value
      cv2 = dl2[t]
      d.append([kh, cv2, 'XS', xsv])
      d.append([kh, cv2, 'S', sv])
      d.append([kh, cv2, 'M', mv])
      d.append([kh, cv2, 'L', lv])
      ad.append([kh, cv2, 'XS', xsv])
      ad.append([kh, cv2, 'S', sv])
      ad.append([kh, cv2, 'M', mv])
      ad.append([kh, cv2, 'L', lv])
     t = t + 1

  # 写入款式表格
  wb2 = Workbook()
  sheet = wb2["Sheet"]
  sheet.cell(row=1, column=1).value = '款式编码'
  sheet.cell(row=1, column=2).value = '部位名称'
  sheet.cell(row=1, column=3).value = '尺码名称'
  sheet.cell(row=1, column=4).value = '尺码数据'
  for i in range(len(d)):
   sheet.cell(row=i+2, column=1).value = d[i][0]
   sheet.cell(row=i+2, column=2).value = d[i][1]
   sheet.cell(row=i+2, column=3).value = d[i][2]
   sheet.cell(row=i+2, column=4).value = d[i][3]
  wb2.save('Excel2/' + kh + '.xlsx')

# 写入汇总表格
wb3 = Workbook()
sheet2 = wb3["Sheet"]
sheet2.cell(row=1, column=1).value = '款式编码'
sheet2.cell(row=1, column=2).value = '部位名称'
sheet2.cell(row=1, column=3).value = '尺码名称'
sheet2.cell(row=1, column=4).value = '尺码数据'
for i1 in range(len(ad)):
 sheet2.cell(row=i1+2, column=1).value = ad[i1][0]
 sheet2.cell(row=i1+2, column=2).value = ad[i1][1]
 sheet2.cell(row=i1+2, column=3).value = ad[i1][2]
 sheet2.cell(row=i1+2, column=4).value = ad[i1][3]
wb3.save('Excel2/all.xlsx')

#
# def print_hi(name):
#     # Use a breakpoint in the code line below to debug your script.
#     print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.
#
#
# # Press the green button in the gutter to run the script.
# if __name__ == '__main__':
#     print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
