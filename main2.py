# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os, re
from openpyxl import load_workbook, Workbook

dl = ['胸围', '肩宽']
dl2 = ['Bust', 'Shoulder']
# 获取当前路径
cur_dir = os.path.dirname(os.path.abspath(__file__))
# 遍历当前路径下的Excel文件夹
d = []

for root, dirs, files in os.walk(cur_dir + '\\Excel'):
 for file in files:
  wb = load_workbook(filename='Excel/' + file, data_only=True)
  sheet_name = wb.sheetnames[3]
  sheet_ranges = wb[sheet_name]
  # 取款号
  kh = sheet_ranges['C2'].value
  print(sheet_ranges['C2'].value)  # 3
  # 循环判断是否是要的度量方法
  for cell in sheet_ranges['B']:
   cv = cell.value
   ci = cell.row
   if cv != None:
    # 判断是否要的度量
    t = 0
    for i in dl:
     if re.match(i, cv) != None:
      xsv = sheet_ranges['D' + str(ci)].value
      sv = sheet_ranges['E' + str(ci)].value
      mv = sheet_ranges['F' + str(ci)].value
      lv = sheet_ranges['G' + str(ci)].value
      cv2 = dl2[t]
      d.append([kh, cv2, 'XS', xsv])
      d.append([kh, cv2, 'S', sv])
      d.append([kh, cv2, 'M', mv])
      d.append([kh, cv2, 'L', lv])
     t = t + 1
     # 循环4次，生成4条记录存起来，格式为：款号，度量方法，数值

  # 写入目标表格
wb2 = Workbook()
sheet = wb2["Sheet"]
sheet.cell(row=1, column=1).value = '款式编码'
sheet.cell(row=1, column=2).value = '部位名称'
sheet.cell(row=1, column=3).value = '尺码名称'
sheet.cell(row=1, column=4).value = '尺码数据'
for i in range(len(d)):
 sheet.cell(row=i+2, column=1).value = d[i][0]
 sheet.cell(row=i+2, column=2).value = d[i][1]
 sheet.cell(row=i+2, column=3).value = d[i][2]
 sheet.cell(row=i+2, column=4).value = d[i][3]
wb2.save('Excel2/all.xlsx')

